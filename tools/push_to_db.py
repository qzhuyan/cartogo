#!/usr/bin/python
# coding=utf-8
from openpyxl import *
from elasticsearch import Elasticsearch

wb = load_workbook('testdat.xlsx')

ws =  wb['Sheet1']

alldata = []
for row in range(3,6000):
    data = {'Client' : ws.cell(column=5, row=row).value,
            'CarId' : ws.cell(column=6, row=row).value,
            'Rent' : ws.cell(column=7, row=row).value,
            'Return' : ws.cell(column=8, row=row).value,
            '@timestamp': ws.cell(column=7, row=row).value}

    alldata.append(data)

DBServerUrl = 'http://192.168.200.8:34567'

es = Elasticsearch([DBServerUrl])

for doc in alldata:
    print es.index(index="testdb1", doc_type='car2go_rent',  body=doc)





    
    


